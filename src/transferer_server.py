"""
Transferer Server - Web interface for inventory transfers
Cross-platform compatible (Windows, macOS, Linux).
"""
from flask import Flask, render_template, request, jsonify, Response, send_file
from openpyxl import load_workbook, Workbook
import io
import json
import subprocess
import threading
import queue
import time
import os
import sys
from pathlib import Path

# Setup cross-platform imports
_script_dir = Path(__file__).parent
_project_root = _script_dir.parent

# Add paths for imports to work from any launch location
sys.path.insert(0, str(_script_dir))
sys.path.insert(0, str(_project_root))

from android_controller import FinaleAutomator
from adb_utils import get_data_file_path, create_stop_signal, clear_stop_signal, PRODUCT_MARKER
from error_detector import ErrorDetector
from screen_inspect import OCR_AVAILABLE

app = Flask(__name__)

# Global state for batch execution
automator = None
execution_queue = queue.Queue()
execution_status = {
    'running': False,
    'current': 0,
    'total': 0,
    'current_item': None,
    'message': '',
    'error_detected': False
}
# Store batch data: locations from row 1, all IMEIs from column C
pending_batch = {
    'from_loc': '',
    'to_loc': '',
    'imeis': []
}

# Change Item State specific state
change_state_status = {
    'running': False,
    'current': 0,
    'total': 0,
    'current_item': None,
    'message': '',
    'error_detected': False
}
pending_change_state_items = []
change_state_process = None  # Store subprocess for stop functionality

# Receive specific state
receive_status = {
    'running': False,
    'current': 0,
    'total': 0,
    'offset': 0,
    'current_item': None,
    'message': '',
    'error_detected': False,
    'skipped': 0,
    'ocr_available': OCR_AVAILABLE
}
pending_receive_items = []
receive_process = None  # Store subprocess for stop functionality

# Pick specific state
pick_status = {
    'running': False,
    'current': 0,
    'total': 0,
    'current_item': None,
    'message': '',
    'error_detected': False
}
pending_pick_items = []
pick_process = None  # Store subprocess for stop functionality


# ============================================================
# DEVICE LOCK - one Android device, one batch at a time
# ============================================================
# Every flow types into the same phone; two running at once interleave
# keystrokes and silently corrupt inventory data.
device_lock = threading.Lock()
device_owner = None  # flow key currently driving the device, or None

FLOW_LABELS = {
    'stock': 'Stock',
    'transfer': 'Transfer',
    'receive': 'Receive',
    'pick': 'Pick',
    'change_state': 'Change Item State',
}


def try_acquire_device(flow):
    """Claim the device for a flow. Returns None on success, else the flow
    key currently holding it."""
    global device_owner
    with device_lock:
        if device_owner is None:
            device_owner = flow
            return None
        return device_owner


def release_device(flow):
    """Release the device if held by this flow (idempotent)."""
    global device_owner
    with device_lock:
        if device_owner == flow:
            device_owner = None


def force_release_device():
    """Clear the device claim regardless of owner (used by /reset)."""
    global device_owner
    with device_lock:
        device_owner = None


def device_busy_response(owner):
    label = FLOW_LABELS.get(owner, owner)
    return jsonify({
        'success': False,
        'message': f'A {label} batch is running. Stop it first.'
    })


def get_automator():
    """Get or create the FinaleAutomator instance"""
    global automator
    if automator is None:
        automator = FinaleAutomator()
    return automator


@app.route('/')
def index():
    """Serve the home page with navigation"""
    return render_template('index.html')


@app.route('/transfer')
def transfer_page():
    """Serve the transfer interface"""
    return render_template('transferer.html')


@app.route('/transfer', methods=['POST'])
def transfer():
    """Handle single transfer request (queue only, no execution)"""
    data = request.json
    from_location = data.get('from_location', '')
    to_location = data.get('to_location', '')
    imei = data.get('imei', '')
    
    if not all([from_location, to_location, imei]):
        return jsonify({'success': False, 'message': 'All fields are required'})
    
    print(f"Transfer queued: {from_location} -> {to_location}, IMEI: {imei}")
    
    return jsonify({
        'success': True,
        'message': f'Transfer queued: {imei} from {from_location} to {to_location}'
    })


@app.route('/execute', methods=['POST'])
def execute_single():
    """Execute a single transfer immediately via ADB"""
    global execution_status
    
    if execution_status['running']:
        return jsonify({'success': False, 'message': 'Another execution is in progress'})
    
    data = request.json
    from_location = data.get('from_location', '')
    to_location = data.get('to_location', '')
    imei = data.get('imei', '')
    
    if not all([from_location, to_location, imei]):
        return jsonify({'success': False, 'message': 'All fields are required'})

    owner = try_acquire_device('transfer')
    if owner:
        return device_busy_response(owner)

    try:
        execution_status['running'] = True
        execution_status['current'] = 1
        execution_status['total'] = 1
        execution_status['message'] = 'Executing transfer...'

        auto = get_automator()
        result = auto.execute_transfer(from_location, to_location, imei)

        execution_status['running'] = False
        execution_status['message'] = result['message']

        return jsonify(result)

    except Exception as e:
        execution_status['running'] = False
        execution_status['message'] = f'Error: {str(e)}'
        return jsonify({'success': False, 'message': str(e)})
    finally:
        release_device('transfer')

# NOTE: Change Item State upload at /change-state/upload writes to receive.txt and runs change_item_state_auto.py
# This /upload route is for Transfer: reads col A=from, B=to, C=imei starting from row 2

# Transfer specific state
transfer_status = {
    'running': False,
    'current': 0,
    'total': 0,
    'current_item': None,
    'message': '',
    'error_detected': False
}
pending_transfer_items = []
transfer_process = None  # Store subprocess for stop functionality


@app.route('/upload', methods=['POST'])
def upload_excel():
    """
    Handle Excel file upload for Transfer.
    
    Excel format:
    - Row 1: Header (ignored)
    - Row 2+: Column A = From sublocation, Column B = To sublocation, Column C = IMEI
    - From/To are read from first data row (row 2) and apply to all IMEIs
    """
    global pending_batch, pending_transfer_items
    
    # Clear error state and stop signal from any previous run
    transfer_status['error_detected'] = False
    clear_stop_signal('transfer')
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'File must be .xlsx or .xls'})
    
    try:
        # Read Excel file
        wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        
        # Read from row 2 (skip header in row 1)
        # Get from/to from first data row, collect all IMEIs from column C
        from_loc = None
        to_loc = None
        imeis = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ensure row has at least 3 columns (A, B, C)
            if not row or len(row) < 3:
                continue
            
            # Get from/to from first data row
            if from_loc is None and row[0] is not None:
                from_loc = str(row[0]).strip()
            if to_loc is None and row[1] is not None:
                to_loc = str(row[1]).strip()
            
            # Collect IMEI from column C
            if row[2] is not None:
                # Handle numeric IMEIs (Excel may read as float)
                imei_val = row[2]
                if isinstance(imei_val, float):
                    imei_val = int(imei_val)
                imei_str = str(imei_val).strip()
                if imei_str:
                    imeis.append(imei_str)
        
        wb.close()
        
        if not from_loc or not to_loc:
            return jsonify({'success': False, 'message': 'Row 2 must have From (A) and To (B) sublocations'})
        
        if not imeis:
            return jsonify({'success': False, 'message': 'No IMEIs found in column C (starting row 2)'})
        
        # Store batch data for execution
        pending_batch = {
            'from_loc': from_loc,
            'to_loc': to_loc,
            'imeis': imeis
        }
        pending_transfer_items = imeis
        
        # Write to transfer_data.txt for transfer_auto.py
        transfer_data_file = get_data_file_path('transfer_data.txt')
        with open(transfer_data_file, 'w') as f:
            f.write(f"{from_loc}\n")
            f.write(f"{to_loc}\n")
            for imei in imeis:
                f.write(f"{imei}\n")
        
        print(f"Loaded transfer batch: {from_loc} -> {to_loc}, {len(imeis)} IMEIs")
        print(f"Saved to {transfer_data_file}")
        for imei in imeis[:5]:
            print(f"  IMEI: {imei}")
        if len(imeis) > 5:
            print(f"  ... and {len(imeis) - 5} more")
        
        # Return in format compatible with frontend
        transfers = [{'from': from_loc, 'to': to_loc, 'imei': imei} for imei in imeis]
        
        return jsonify({
            'success': True,
            'message': f'Loaded {len(imeis)} IMEIs ({from_loc} → {to_loc})',
            'transfers': transfers
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error reading file: {str(e)}'})


def execute_batch_worker(from_loc, to_loc, imeis):
    """Background worker for batch execution - same locations for all IMEIs"""
    global execution_status

    try:
        auto = get_automator()

        def progress_callback(current, total, status, imei):
            execution_status['current'] = current
            execution_status['total'] = total
            execution_status['current_item'] = {'from': from_loc, 'to': to_loc, 'imei': imei}
            execution_status['message'] = f'{status}: {imei}'

        result = auto.execute_same_location_batch(from_loc, to_loc, imeis, progress_callback)

        execution_status['message'] = result['message']
        execution_status['result'] = result
    except Exception as e:
        execution_status['message'] = f'Error: {str(e)}'
        execution_status['result'] = {'success': False, 'message': str(e)}
    finally:
        execution_status['running'] = False
        release_device('stock')


@app.route('/execute-batch', methods=['POST'])
def execute_batch():
    """Start batch execution of pending transfers"""
    global execution_status, pending_batch
    
    if execution_status['running']:
        return jsonify({'success': False, 'message': 'Another execution is in progress'})
    
    # Use pending batch data
    from_loc = pending_batch.get('from_loc', '')
    to_loc = pending_batch.get('to_loc', '')
    imeis = pending_batch.get('imeis', [])
    
    if not imeis:
        return jsonify({'success': False, 'message': 'No IMEIs to execute'})
    
    if not from_loc or not to_loc:
        return jsonify({'success': False, 'message': 'Missing from/to locations'})

    owner = try_acquire_device('stock')
    if owner:
        return device_busy_response(owner)

    # Reset status
    execution_status = {
        'running': True,
        'current': 0,
        'total': len(imeis),
        'current_item': None,
        'message': 'Starting batch execution...',
        'result': None
    }
    
    # Start background thread
    thread = threading.Thread(target=execute_batch_worker, args=(from_loc, to_loc, imeis))
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'message': f'Started batch: {len(imeis)} IMEIs ({from_loc} → {to_loc})'
    })


def read_progress_file(filename):
    """Read progress from a progress file. Returns (current, total) or None."""
    try:
        progress_file = get_data_file_path(filename)
        if os.path.exists(progress_file):
            with open(progress_file, 'r') as f:
                content = f.read().strip()
                if ',' in content:
                    current, total = content.split(',')
                    return int(current), int(total)
    except:
        pass
    return None


def compute_resume_remainder(flow):
    """Build trimmed data-file content so a stopped batch continues where it
    left off instead of re-running completed items.

    Reads the flow's progress file (current,total) and data file. Returns
    {'content': str, 'done': int, 'remaining': int} or None when resume is
    not possible (no/finished progress, or files inconsistent with the run).
    """
    progress_files = {
        'transfer': 'transfer_progress.txt',
        'pick': 'pick_progress.txt',
        'receive': 'receive_progress.txt',
        'change_state': 'change_state_progress.txt',
    }
    data_files = {
        'transfer': 'transfer_data.txt',
        'pick': 'pick_data.txt',
        'receive': 'receive_data.txt',
        'change_state': 'receive.txt',
    }

    progress = read_progress_file(progress_files[flow])
    if not progress:
        return None
    done, total = progress
    if done <= 0 or done >= total:
        return None

    data_path = get_data_file_path(data_files[flow])
    if not os.path.exists(data_path):
        return None

    if flow == 'pick':
        # Raw lines; empty lines are meaningful (ENTER-only). The script drops
        # trailing empties, so mirror that before trimming.
        with open(data_path, 'r') as f:
            lines = [line.rstrip('\n\r') for line in f]
        while lines and lines[-1] == '':
            lines.pop()
        if len(lines) != total:
            return None
        remaining = lines[done:]
        if not remaining:
            return None
        return {
            'content': '\n'.join(remaining) + '\n',
            'done': done,
            'remaining': len(remaining),
        }

    with open(data_path, 'r') as f:
        lines = [line.strip() for line in f if line.strip()]

    if flow == 'transfer':
        # Line 1 from, line 2 to, rest IMEIs; progress counts IMEIs.
        if len(lines) < 3 or len(lines) - 2 != total:
            return None
        header, imeis = lines[:2], lines[2:]
        remaining = imeis[done:]
        if not remaining:
            return None
        return {
            'content': '\n'.join(header + remaining) + '\n',
            'done': done,
            'remaining': len(remaining),
        }

    if flow == 'receive':
        # Marked product/IMEI lines; progress counts all lines. Only resume
        # marker-format files (anything else can't be trimmed safely).
        if len(lines) != total or not any(l.startswith(PRODUCT_MARKER) for l in lines):
            return None
        remaining = lines[done:]
        if not remaining:
            return None
        # Re-establish product context when resuming mid-group.
        if not remaining[0].startswith(PRODUCT_MARKER):
            last_product = None
            for l in lines[:done]:
                if l.startswith(PRODUCT_MARKER):
                    last_product = l
            if last_product:
                remaining = [last_product] + remaining
        return {
            'content': '\n'.join(remaining) + '\n',
            'done': done,
            'remaining': len(remaining),
        }

    if flow == 'change_state':
        # IMEI/ProductID pairs; progress counts pairs.
        pairs = [(lines[i], lines[i + 1]) for i in range(0, len(lines) - 1, 2)]
        if len(pairs) != total:
            return None
        remaining = pairs[done:]
        if not remaining:
            return None
        flat = [v for pair in remaining for v in pair]
        return {
            'content': '\n'.join(flat) + '\n',
            'done': done,
            'remaining': len(remaining),
        }

    return None


def apply_resume(flow, status_dict):
    """Trim the flow's data file for a resume request.

    Returns (offset, remaining) when the batch will continue from where it
    stopped, or None when resume isn't possible (caller falls back to a full
    run). Offsets accumulate across chained resumes for correct display.
    """
    trimmed = compute_resume_remainder(flow)
    if not trimmed:
        return None
    data_files = {
        'transfer': 'transfer_data.txt',
        'pick': 'pick_data.txt',
        'receive': 'receive_data.txt',
        'change_state': 'receive.txt',
    }
    progress_files = {
        'transfer': 'transfer_progress.txt',
        'pick': 'pick_progress.txt',
        'receive': 'receive_progress.txt',
        'change_state': 'change_state_progress.txt',
    }
    with open(get_data_file_path(data_files[flow]), 'w') as f:
        f.write(trimmed['content'])
    with open(get_data_file_path(progress_files[flow]), 'w') as f:
        f.write(f"0,{trimmed['remaining']}")
    offset = status_dict.get('offset', 0) + trimmed['done']
    print(f"Resume {flow}: {trimmed['done']} done this segment "
          f"({offset} total), {trimmed['remaining']} remaining")
    return offset, trimmed['remaining']


def count_skipped_file(filename='receive_skipped.txt'):
    """Count non-empty lines in a skipped-items log file (0 if missing)."""
    try:
        path = get_data_file_path(filename)
        if os.path.exists(path):
            with open(path, 'r') as f:
                return sum(1 for line in f if line.strip())
    except:
        pass
    return 0


def execute_transfer_batch_worker(total):
    """Background worker for batch Transfer execution - runs transfer_auto.py"""
    global transfer_status, transfer_process

    script_path = str(_project_root / 'transfer_auto.py')
    progress_file = 'transfer_progress.txt'
    error_detector = None
    error_at_item = [None]  # Use list to allow modification in nested function

    def on_error_detected():
        """Callback when error detector finds red screen"""
        # Capture the current progress at the moment of error
        error_at_item[0] = transfer_status['current']
        transfer_status['error_detected'] = True
        transfer_status['message'] = ('Error detected: Red screen - stopping. '
             'Dismiss the error on the device, then press Resume.')
        # Create stop signal FIRST so script sees it before terminate
        create_stop_signal('transfer')
        if transfer_process:
            try:
                transfer_process.terminate()
            except:
                pass

    def on_crash_detected():
        """Callback when app crash/ANR is detected"""
        error_at_item[0] = transfer_status['current']
        transfer_status['error_detected'] = True
        transfer_status['message'] = (
            'App crashed: Finale app stopped responding. '
            'Dialog dismissed and app reopened. Return to the correct page and press Resume.'
        )
        create_stop_signal('transfer')
        if transfer_process:
            try:
                transfer_process.terminate()
            except:
                pass

    try:
        # Clear any existing stop signal before starting
        clear_stop_signal('transfer')
        transfer_status['current'] = 0
        transfer_status['total'] = total
        transfer_status['message'] = 'Starting transfer_auto.py...'

        # Start error detector
        error_detector = ErrorDetector(
            callback=on_error_detected,
            crash_callback=on_crash_detected
        )
        error_detector.start()

        # Start the process (non-blocking)
        transfer_process = subprocess.Popen(
            [sys.executable, script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            cwd=str(_project_root)
        )

        # Poll for progress while process is running
        while transfer_process.poll() is None:
            # Check if error was detected - stop updating progress
            if transfer_status['error_detected']:
                transfer_process.terminate()
                break

            progress = read_progress_file(progress_file)
            if progress:
                current, _ = progress
                # Only update progress if no error detected
                if not transfer_status['error_detected']:
                    transfer_status['current'] = current
                    transfer_status['message'] = f'Processing {current}/{total}...'
            time.sleep(0.3)
        
        # Process finished
        stdout, stderr = transfer_process.communicate()
        stderr_text = stderr.decode().strip() if stderr else ''

        if transfer_status.get('error_detected', False):
            # Red screen or crash was detected by ErrorDetector
            completed = error_at_item[0] if error_at_item[0] is not None else transfer_status['current']
            transfer_status['current'] = completed
            transfer_status['result'] = {
                'success': False,
                'completed': completed,
                'total': total,
                'message': transfer_status['message']  # Keep the error message
            }
            transfer_status['running'] = False
        elif transfer_process.returncode == 0:
            transfer_status['current'] = total
            transfer_status['message'] = f'Completed all {total} transfers'
            transfer_status['result'] = {
                'success': True,
                'completed': total,
                'total': total,
                'message': f'Completed all {total} transfers'
            }
            transfer_status['running'] = False  # Set last to avoid race condition
        elif stderr_text:
            # Real script error with diagnostic output
            transfer_status['running'] = False
            transfer_status['message'] = f'Script error: {stderr_text}'
            transfer_status['result'] = {
                'success': False,
                'completed': transfer_status['current'],
                'total': total,
                'message': f'Script error: {stderr_text}'
            }
        else:
            # Non-zero exit with no stderr: process was terminated (user stop
            # or Windows terminate()=1). Clear status silently.
            transfer_status['message'] = ''
            transfer_status['result'] = {
                'success': False,
                'completed': transfer_status['current'],
                'total': total,
                'message': ''
            }
            transfer_status['running'] = False
            
    except Exception as e:
        transfer_status['running'] = False
        transfer_status['message'] = f'Error: {str(e)}'
        transfer_status['result'] = {
            'success': False,
            'completed': 0,
            'total': total,
            'message': f'Error: {str(e)}'
        }
    finally:
        # Stop error detector
        if error_detector:
            error_detector.stop()
        transfer_process = None
        # The stop signal only targets the run that just ended; never leave it
        # behind to kill a future run.
        clear_stop_signal('transfer')
        release_device('transfer')


@app.route('/execute-transfer-batch', methods=['POST'])
def execute_transfer_batch():
    """Start batch execution of Transfer operations using transfer_auto.py"""
    global transfer_status, pending_transfer_items
    
    if transfer_status['running']:
        return jsonify({'success': False, 'message': 'Another execution is in progress'})

    owner = try_acquire_device('transfer')
    if owner:
        return device_busy_response(owner)

    data = request.json or {}
    offset = 0
    total = len(pending_transfer_items)
    resumed = apply_resume('transfer', transfer_status) if data.get('resume') else None
    if resumed:
        offset, total = resumed
    elif not pending_transfer_items:
        release_device('transfer')
        return jsonify({'success': False, 'message': 'No items to execute. Upload Excel file first.'})

    # Reset status
    transfer_status = {
        'running': True,
        'current': 0,
        'total': total,
        'offset': offset,
        'current_item': None,
        'message': f'Resuming: {offset} already done...' if resumed else 'Starting batch execution...',
        'result': None,
        'error_detected': False
    }
    
    # Start background thread to run transfer_auto.py
    thread = threading.Thread(target=execute_transfer_batch_worker, args=(total,))
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'message': f'Started batch: {total} transfers'
    })


@app.route('/transfer-status-stream')
def transfer_status_stream():
    """Server-Sent Events stream for real-time Transfer status updates"""
    def generate():
        last_status = None
        while True:
            current = json.dumps(transfer_status)
            if current != last_status:
                yield f"data: {current}\n\n"
                last_status = current
            time.sleep(0.3)  # Poll every 300ms
            
            # Stop streaming if not running and we've sent the final status
            if not transfer_status['running'] and last_status == current:
                yield f"data: {current}\n\n"
                break
    
    return Response(generate(), mimetype='text/event-stream')


@app.route('/stop', methods=['POST'])
def stop_execution():
    """Stop the current batch execution (supports both automator and script-based)"""
    global execution_status, transfer_status, transfer_process
    
    # Check if transfer script is running
    if transfer_status['running'] and transfer_process is not None:
        try:
            transfer_process.terminate()
            transfer_status['message'] = 'Stopping...'
            return jsonify({'success': True, 'message': 'Stop requested'})
        except:
            pass
    
    # Check if automator-based execution is running
    if execution_status['running']:
        auto = get_automator()
        auto.request_stop()
        return jsonify({'success': True, 'message': 'Stop requested'})
    
    return jsonify({'success': False, 'message': 'No execution in progress'})


@app.route('/status')
def get_status():
    """Get current execution status"""
    return jsonify(execution_status)


@app.route('/status-stream')
def status_stream():
    """Server-Sent Events stream for real-time status updates"""
    def generate():
        last_status = None
        while True:
            current = json.dumps(execution_status)
            if current != last_status:
                yield f"data: {current}\n\n"
                last_status = current
            time.sleep(0.3)  # Poll every 300ms
            
            # Stop streaming if not running and we've sent the final status
            if not execution_status['running'] and last_status == current:
                # Send one more to ensure client gets final state
                yield f"data: {current}\n\n"
                break
    
    return Response(generate(), mimetype='text/event-stream')


@app.route('/device-status')
def device_status():
    """Check if Android device is connected"""
    try:
        auto = get_automator()
        devices = auto.get_devices()
        connected = any(d['status'] == 'device' for d in devices)
        return jsonify({
            'connected': connected,
            'devices': devices
        })
    except Exception as e:
        return jsonify({
            'connected': False,
            'devices': [],
            'error': str(e)
        })


@app.route('/active-batch')
def active_batch():
    """Which flow currently owns the device (if any) and its status snapshot.

    Lets pages reconnect to a running batch after a refresh and warn when a
    different flow holds the device.
    """
    flow = device_owner
    statuses = {
        'stock': execution_status,
        'transfer': transfer_status,
        'receive': receive_status,
        'pick': pick_status,
        'change_state': change_state_status,
    }
    return jsonify({
        'flow': flow,
        'label': FLOW_LABELS.get(flow),
        'status': statuses.get(flow)
    })


@app.route('/bulk-stock')
def bulk_stock_page():
    """Serve the Bulk Stock interface"""
    return render_template('bulkStock.html')


# ============================================================
# CHANGE ITEM STATE ROUTES
# ============================================================

@app.route('/change-state')
def change_state_page():
    """Serve the Change Item State interface"""
    return render_template('changeItemState.html')

@app.route('/demo')
def demo_page():
    """Serve the demo page"""
    return render_template('demo.html')


@app.route('/change-state/upload', methods=['POST'])
def upload_change_state_excel():
    """
    Handle Excel file upload for Change Item State.
    
    Excel format:
    - Column A: IMEI (starting from row 2)
    - Column B: New Product ID (starting from row 2)
    - Row 1 is header row (ignored)
    """
    global pending_change_state_items
    
    # Clear error state and stop signal from any previous run
    change_state_status['error_detected'] = False
    clear_stop_signal('change_state')
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'File must be .xlsx or .xls'})
    
    try:
        # Read Excel file
        wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        
        # Collect IMEI (column A) and Product ID (column B) pairs
        # Starting from row 2 (row 1 is header)
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ensure row has at least 2 columns
            if not row or len(row) < 2:
                continue
            
            # Column A = IMEI, Column B = Product ID
            imei_val = row[0]
            product_id_val = row[1]
            
            # Skip if either value is empty
            if imei_val is None or product_id_val is None:
                continue
            
            # Handle numeric IMEIs (Excel may read as float)
            if isinstance(imei_val, float):
                imei_val = int(imei_val)
            imei_str = str(imei_val).strip()
            product_id_str = str(product_id_val).strip()
            
            if imei_str and product_id_str:
                items.append({
                    'imei': imei_str,
                    'new_product_id': product_id_str
                })
        
        wb.close()
        
        if not items:
            return jsonify({'success': False, 'message': 'No valid IMEI/Product ID pairs found (check columns A and B, starting from row 2)'})
        
        # Store items for execution
        pending_change_state_items = items
        
        # Write to receive.txt in format expected by change_item_state_auto.py
        # Format: imei\nproductID\nimei\nproductID\n...
        receive_file = get_data_file_path('receive.txt')
        with open(receive_file, 'w') as f:
            for item in items:
                f.write(f"{item['imei']}\n")
                f.write(f"{item['new_product_id']}\n")
        
        print(f"Loaded {len(items)} Change Item State pairs and wrote to receive.txt:")
        for item in items[:5]:  # Show first 5
            print(f"  {item['imei']} -> {item['new_product_id']}")
        if len(items) > 5:
            print(f"  ... and {len(items) - 5} more")
        
        return jsonify({
            'success': True,
            'message': f'Loaded {len(items)} IMEI/Product ID pairs (saved to receive.txt)',
            'items': items
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error reading file: {str(e)}'})


@app.route('/change-state/execute', methods=['POST'])
def execute_single_change_state():
    """Execute a single Change Item State operation - writes to receive.txt and runs script"""
    global change_state_status
    
    if change_state_status['running']:
        return jsonify({'success': False, 'message': 'Another execution is in progress'})
    
    data = request.json
    imei = data.get('imei', '')
    new_product_id = data.get('new_product_id', '')
    
    if not imei or not new_product_id:
        return jsonify({'success': False, 'message': 'IMEI and Product ID are required'})

    owner = try_acquire_device('change_state')
    if owner:
        return device_busy_response(owner)

    try:
        change_state_status['running'] = True
        change_state_status['current'] = 1
        change_state_status['total'] = 1
        change_state_status['message'] = 'running change_item_state_auto.py'
        
        # Write single item to receive.txt
        receive_file = get_data_file_path('receive.txt')
        with open(receive_file, 'w') as f:
            f.write(f"{imei}\n")
            f.write(f"{new_product_id}\n")
        
        # Run change_item_state_auto.py
        script_path = str(_project_root / 'change_item_state_auto.py')
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            cwd=str(_project_root)
        )
        
        change_state_status['running'] = False
        
        if result.returncode == 0:
            change_state_status['message'] = f'Changed {imei} to {new_product_id}'
            return jsonify({'success': True, 'message': f'Changed {imei} to {new_product_id}'})
        else:
            change_state_status['message'] = f'Script error: {result.stderr}'
            return jsonify({'success': False, 'message': f'Script error: {result.stderr}'})
        
    except Exception as e:
        change_state_status['running'] = False
        change_state_status['message'] = f'Error: {str(e)}'
        return jsonify({'success': False, 'message': str(e)})
    finally:
        release_device('change_state')


def execute_change_state_batch_worker(items, total=None):
    """Background worker for batch Change Item State execution - runs change_item_state_auto.py"""
    global change_state_status, change_state_process

    total = total if total is not None else len(items)
    script_path = str(_project_root / 'change_item_state_auto.py')
    progress_file = 'change_state_progress.txt'
    error_detector = None
    error_at_item = [None]  # Use list to allow modification in nested function

    def on_error_detected():
        """Callback when error detector finds red screen"""
        # Capture the current progress at the moment of error
        error_at_item[0] = change_state_status['current']
        change_state_status['error_detected'] = True
        change_state_status['message'] = ('Error detected: Red screen - stopping. '
             'Dismiss the error on the device, then press Resume.')
        # Create stop signal FIRST so script sees it before terminate
        create_stop_signal('change_state')
        if change_state_process:
            try:
                change_state_process.terminate()
            except:
                pass

    def on_crash_detected():
        """Callback when app crash/ANR is detected"""
        error_at_item[0] = change_state_status['current']
        change_state_status['error_detected'] = True
        change_state_status['message'] = (
            'App crashed: Finale app stopped responding. '
            'Dialog dismissed and app reopened. Return to the correct page and press Resume.'
        )
        create_stop_signal('change_state')
        if change_state_process:
            try:
                change_state_process.terminate()
            except:
                pass

    try:
        # Clear any existing stop signal before starting
        clear_stop_signal('change_state')
        change_state_status['current'] = 0
        change_state_status['total'] = total
        change_state_status['message'] = 'Starting change_item_state_auto.py...'

        # Start error detector
        error_detector = ErrorDetector(
            callback=on_error_detected,
            crash_callback=on_crash_detected
        )
        error_detector.start()

        # Start the process (non-blocking)
        change_state_process = subprocess.Popen(
            [sys.executable, script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            cwd=str(_project_root)
        )

        # Poll for progress while process is running
        while change_state_process.poll() is None:
            # Check if error was detected - stop updating progress
            if change_state_status['error_detected']:
                change_state_process.terminate()
                break

            progress = read_progress_file(progress_file)
            if progress:
                current, _ = progress
                # Only update progress if no error detected
                if not change_state_status['error_detected']:
                    change_state_status['current'] = current
                    change_state_status['message'] = f'Processing {current}/{total}...'
            time.sleep(0.3)
        
        # Process finished
        stdout, stderr = change_state_process.communicate()
        stderr_text = stderr.decode().strip() if stderr else ''

        if change_state_status.get('error_detected', False):
            # Red screen or crash was detected by ErrorDetector
            completed = error_at_item[0] if error_at_item[0] is not None else change_state_status['current']
            change_state_status['current'] = completed
            change_state_status['result'] = {
                'success': False,
                'completed': completed,
                'total': total,
                'message': change_state_status['message']  # Keep the error message
            }
            change_state_status['running'] = False
        elif change_state_process.returncode == 0:
            change_state_status['current'] = total
            change_state_status['message'] = f'Completed all {total} items'
            change_state_status['result'] = {
                'success': True,
                'completed': total,
                'total': total,
                'message': f'Completed all {total} item state changes'
            }
            change_state_status['running'] = False  # Set last to avoid race condition
        elif stderr_text:
            # Real script error with diagnostic output
            change_state_status['running'] = False
            change_state_status['message'] = f'Script error: {stderr_text}'
            change_state_status['result'] = {
                'success': False,
                'completed': change_state_status['current'],
                'total': total,
                'message': f'Script error: {stderr_text}'
            }
        else:
            # Non-zero exit with no stderr: process was terminated (user stop
            # or Windows terminate()=1). Clear status silently.
            change_state_status['message'] = ''
            change_state_status['result'] = {
                'success': False,
                'completed': change_state_status['current'],
                'total': total,
                'message': ''
            }
            change_state_status['running'] = False
            
    except Exception as e:
        change_state_status['running'] = False
        change_state_status['message'] = f'Error: {str(e)}'
        change_state_status['result'] = {
            'success': False,
            'completed': 0,
            'total': total,
            'message': f'Error: {str(e)}'
        }
    finally:
        # Stop error detector
        if error_detector:
            error_detector.stop()
        change_state_process = None
        clear_stop_signal('change_state')
        release_device('change_state')


@app.route('/change-state/execute-batch', methods=['POST'])
def execute_change_state_batch():
    """Start batch execution of Change Item State"""
    global change_state_status, pending_change_state_items
    
    if change_state_status['running']:
        return jsonify({'success': False, 'message': 'Another execution is in progress'})

    owner = try_acquire_device('change_state')
    if owner:
        return device_busy_response(owner)

    # Use items from request or pending items
    data = request.json or {}
    items = data.get('items', pending_change_state_items)

    offset = 0
    total = len(items)
    resumed = apply_resume('change_state', change_state_status) if data.get('resume') else None
    if resumed:
        offset, total = resumed
    elif not items:
        release_device('change_state')
        return jsonify({'success': False, 'message': 'No items to execute'})

    # Reset status
    change_state_status = {
        'running': True,
        'current': 0,
        'total': total,
        'offset': offset,
        'current_item': None,
        'message': f'Resuming: {offset} already done...' if resumed else 'Starting batch execution...',
        'result': None,
        'error_detected': False
    }

    # Start background thread
    thread = threading.Thread(target=execute_change_state_batch_worker, args=(items, total))
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'message': f'Started batch: {len(items)} items'
    })


@app.route('/change-state/stop', methods=['POST'])
def stop_change_state():
    """Stop the current Change Item State batch execution"""
    global change_state_status, change_state_process
    
    if not change_state_status['running']:
        return jsonify({'success': False, 'message': 'No execution in progress'})
    
    # Terminate the subprocess if running
    if change_state_process is not None:
        try:
            change_state_process.terminate()
            change_state_status['message'] = 'Stopping...'
        except:
            pass
    
    return jsonify({'success': True, 'message': 'Stop requested'})


@app.route('/change-state/status')
def get_change_state_status():
    """Get current Change Item State execution status"""
    return jsonify(change_state_status)


@app.route('/change-state/status-stream')
def change_state_status_stream():
    """Server-Sent Events stream for real-time Change Item State status updates"""
    def generate():
        last_status = None
        while True:
            current = json.dumps(change_state_status)
            if current != last_status:
                yield f"data: {current}\n\n"
                last_status = current
            time.sleep(0.3)  # Poll every 300ms
            
            # Stop streaming if not running and we've sent the final status
            if not change_state_status['running'] and last_status == current:
                yield f"data: {current}\n\n"
                break
    
    return Response(generate(), mimetype='text/event-stream')


# ============================================================
# RECEIVE ROUTES
# ============================================================

@app.route('/receive')
def receive_page():
    """Serve the Receive interface"""
    return render_template('receive.html')


@app.route('/receive', methods=['POST'])
def execute_single_receive():
    """Execute a single receive operation"""
    data = request.json
    order_id = data.get('order_id', '')
    sublocation = data.get('sublocation', '')
    product_id = data.get('product_id', '')
    imei = data.get('imei', '')
    
    if not all([order_id, sublocation, product_id, imei]):
        return jsonify({'success': False, 'message': 'All fields are required'})
    
    owner = try_acquire_device('receive')
    if owner:
        return device_busy_response(owner)

    try:
        # For single receive, we can write to receive_data.txt and execute receive_typing.py
        # Format: product name (from product_id, marked), then IMEI
        receive_data_file = get_data_file_path('receive_data.txt')
        with open(receive_data_file, 'w') as f:
            f.write(f"{PRODUCT_MARKER}{product_id}\n")
            f.write(f"{imei}\n")

        # Execute receive_typing.py in background; hold the device claim
        # until the script finishes.
        script_path = str(_project_root / 'receive_typing.py')
        proc = subprocess.Popen(
            [sys.executable, script_path],
            cwd=str(_project_root)
        )

        def _release_when_done(p):
            p.wait()
            release_device('receive')
        threading.Thread(target=_release_when_done, args=(proc,), daemon=True).start()

        return jsonify({
            'success': True,
            'message': f'Received {imei} for {product_id} (executing receive_typing.py)'
        })

    except Exception as e:
        release_device('receive')
        return jsonify({'success': False, 'message': str(e)})


@app.route('/upload-receive', methods=['POST'])
def upload_receive_excel():
    """
    Handle Excel file upload for Receive.
    
    Excel format:
    - Fixed start: D7
    - QTY header: E6
    - Data rows: Starting from row 7
    - Product name: B7+C7 (concatenated)
    - IMEIs: Column D (D7, D8, D9, etc.)
    - Quantity: Column E (E7, E8, etc.)
    - When E7 contains quantity N (e.g., 5), it means:
      - Product name from B7+C7 applies to rows 7 through (7+N-1) = rows 7-11
      - IMEIs are in D7, D8, D9, D10, D11
      - Next item starts when a new quantity appears in column E (e.g., E12)
    
    Returns products for user to assign custom Product IDs before saving.
    """
    global pending_receive_items
    
    # Clear error state and stop signal from any previous run
    receive_status['error_detected'] = False
    clear_stop_signal('receive')
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'File must be .xlsx or .xls'})
    
    try:
        # Read Excel file
        wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        
        # Pre-load all rows into a list for fast random access (read_only mode)
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
        max_row = len(all_rows)
        
        # Helper to get cell value by 1-based row and column
        def cell_val(r, c):
            if r < 1 or r > max_row:
                return None
            row_data = all_rows[r - 1]
            if c < 1 or c > len(row_data):
                return None
            return row_data[c - 1]
        
        # Parse Excel according to specification
        items = []
        row = 7  # Fixed start at row 7
        
        while row <= max_row:
            # Check quantity cell in column E
            qty_cell = cell_val(row, 5)
            
            # If quantity cell is empty, we're done
            if qty_cell is None or qty_cell == '':
                break
            
            # Convert quantity to int
            try:
                qty = int(qty_cell)
            except (ValueError, TypeError):
                # If not a valid number, skip this row
                row += 1
                continue
            
            # Get product name from B+C (with space between, dashes replaced by spaces)
            b_val = str(cell_val(row, 2) or '').strip()
            c_val = str(cell_val(row, 3) or '').strip()
            product_name = (b_val + ' ' + c_val).strip().replace('-', ' ')
            # Collapse any multiple spaces into one
            product_name = ' '.join(product_name.split())
            
            if not product_name:
                # Skip if no product name
                row += qty
                continue
            
            # Collect IMEIs from column D for the quantity range
            imeis = []
            for i in range(qty):
                imei_cell = cell_val(row + i, 4)
                if imei_cell is not None:
                    # Handle numeric IMEIs (Excel may read as float)
                    if isinstance(imei_cell, float):
                        imei_cell = int(imei_cell)
                    imei_str = str(imei_cell).strip()
                    if imei_str:
                        imeis.append(imei_str)
            
            if imeis:
                items.append({
                    'product_name': product_name,
                    'imeis': imeis
                })
            
            # Move to next quantity group
            row += qty
        
        if not items:
            return jsonify({'success': False, 'message': 'No valid items found in Excel file (check format: QTY in column E starting at row 7, IMEIs in column D, product name in B+C)'})
        
        # Store items for later (user needs to confirm product IDs first)
        pending_receive_items = items
        
        total_imeis = sum(len(item['imeis']) for item in items)
        print(f"Loaded {len(items)} products with {total_imeis} IMEIs (awaiting product ID confirmation):")
        for item in items[:3]:
            print(f"  {item['product_name']}: {len(item['imeis'])} IMEIs")
        if len(items) > 3:
            print(f"  ... and {len(items) - 3} more products")
        
        # Return products for user to assign custom Product IDs
        # Do NOT write to receive_data.txt yet - wait for user confirmation
        products_for_mapping = []
        for i, item in enumerate(items):
            products_for_mapping.append({
                'index': i,
                'detected_name': item['product_name'],
                'imei_count': len(item['imeis']),
                'custom_product_id': ''  # User will fill this in
            })
        
        return jsonify({
            'success': True,
            'message': f'Loaded {len(items)} products with {total_imeis} IMEIs',
            'needs_product_mapping': True,
            'products': products_for_mapping,
            'total_imeis': total_imeis
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error reading file: {str(e)}'})


@app.route('/confirm-receive-products', methods=['POST'])
def confirm_receive_products():
    """
    Confirm product ID mappings and save to receive_data.txt.
    
    Expects JSON:
    {
        "products": [
            {"index": 0, "custom_product_id": "IPHONE 14 128GB BLACK"},
            {"index": 1, "custom_product_id": "IPHONE 15 256GB WHITE"},
            ...
        ]
    }
    """
    global pending_receive_items
    
    if not pending_receive_items:
        return jsonify({'success': False, 'message': 'No pending items. Upload Excel file first.'})
    
    data = request.json
    if not data or 'products' not in data:
        return jsonify({'success': False, 'message': 'Missing product mappings'})
    
    product_mappings = data['products']
    
    # Validate all products have a custom_product_id
    for mapping in product_mappings:
        idx = mapping.get('index')
        custom_id = mapping.get('custom_product_id', '').strip()
        
        if idx is None or idx >= len(pending_receive_items):
            return jsonify({'success': False, 'message': f'Invalid product index: {idx}'})
        
        if not custom_id:
            return jsonify({'success': False, 'message': f'Product {idx + 1} is missing a Product ID'})
    
    # Build the final items with user-provided product IDs
    final_items = []
    for mapping in product_mappings:
        idx = mapping['index']
        custom_id = mapping['custom_product_id'].strip()
        original_item = pending_receive_items[idx]
        
        final_items.append({
            'product_name': custom_id,  # Use user's custom product ID
            'imeis': original_item['imeis']
        })
    
    # Write to receive_data.txt with user-provided product IDs. Product lines
    # are marked so receive_typing.py can tell them from IMEIs without guessing.
    receive_data_file = get_data_file_path('receive_data.txt')
    with open(receive_data_file, 'w') as f:
        for item in final_items:
            f.write(f"{PRODUCT_MARKER}{item['product_name']}\n")
            for imei in item['imeis']:
                f.write(f"{imei}\n")
    
    total_imeis = sum(len(item['imeis']) for item in final_items)
    print(f"Saved {len(final_items)} products with {total_imeis} IMEIs to receive_data.txt:")
    for item in final_items:
        print(f"  {item['product_name']}: {len(item['imeis'])} IMEIs")
    
    # Update pending_receive_items with final product IDs
    pending_receive_items = final_items
    
    # Prepare flattened items list for frontend display
    display_items = []
    for item in final_items:
        for imei in item['imeis']:
            display_items.append({
                'imei': imei,
                'product_name': item['product_name']
            })
    
    return jsonify({
        'success': True,
        'message': f'Saved {len(final_items)} products with {total_imeis} IMEIs',
        'items': display_items
    })


def execute_receive_batch_worker(items, total=None, fresh_skip_log=True):
    """Background worker for batch Receive execution - runs receive_typing.py"""
    global receive_status, receive_process

    total = total if total is not None else len(items)
    script_path = str(_project_root / 'receive_typing.py')
    progress_file = 'receive_progress.txt'
    error_detector = None
    error_at_item = [None]  # Use list to allow modification in nested function

    def on_error_detected():
        """Callback when error detector finds red screen"""
        # Capture the current progress at the moment of error
        # The item that failed is the one being processed (current + 1), so we keep current as-is
        # This represents the last successfully completed item
        error_at_item[0] = receive_status['current']
        receive_status['error_detected'] = True
        receive_status['message'] = ('Error detected: Red screen - stopping. '
             'Dismiss the error on the device, then press Resume.')
        # Create stop signal FIRST so script sees it before terminate
        create_stop_signal('receive')
        if receive_process:
            try:
                receive_process.terminate()
            except:
                pass

    def on_crash_detected():
        """Callback when app crash/ANR is detected"""
        error_at_item[0] = receive_status['current']
        receive_status['error_detected'] = True
        receive_status['message'] = (
            'App crashed: Finale app stopped responding. '
            'Dialog dismissed and app reopened. Return to the correct page and press Resume.'
        )
        create_stop_signal('receive')
        if receive_process:
            try:
                receive_process.terminate()
            except:
                pass

    try:
        # Clear any existing stop signal before starting
        clear_stop_signal('receive')
        # Fresh runs start a new skipped log; resumed runs keep accumulating
        if fresh_skip_log:
            skipped_path = get_data_file_path('receive_skipped.txt')
            try:
                if os.path.exists(skipped_path):
                    os.remove(skipped_path)
            except OSError:
                pass
            receive_status['skipped'] = 0
        receive_status['current'] = 0
        receive_status['total'] = total
        receive_status['message'] = 'Starting receive_typing.py...'

        # Start error detector. detect_red=False: the duplicate-barcode red screen
        # is now handled in-script by receive_typing.py (press Back + skip), so the
        # server must not treat that red screen as a fatal stop. Crash/ANR
        # detection stays on.
        error_detector = ErrorDetector(
            crash_callback=on_crash_detected,
            detect_red=False
        )
        error_detector.start()

        # Start the process (non-blocking)
        receive_process = subprocess.Popen(
            [sys.executable, script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            cwd=str(_project_root)
        )

        # Poll for progress while process is running
        while receive_process.poll() is None:
            # Check if error was detected - stop updating progress
            if receive_status['error_detected']:
                receive_process.terminate()
                break

            progress = read_progress_file(progress_file)
            if progress:
                current, _ = progress
                # Only update progress if no error detected
                if not receive_status['error_detected']:
                    receive_status['current'] = current
                    receive_status['message'] = f'Processing {current}/{total}...'
            # Reflect any IMEIs the script skipped as duplicates
            if not receive_status['error_detected']:
                receive_status['skipped'] = count_skipped_file()
            time.sleep(0.3)
        
        # Process finished
        stdout, stderr = receive_process.communicate()
        stderr_text = stderr.decode().strip() if stderr else ''

        if receive_status.get('error_detected', False):
            # Red screen or crash was detected by ErrorDetector
            completed = error_at_item[0] if error_at_item[0] is not None else receive_status['current']
            receive_status['current'] = completed
            receive_status['result'] = {
                'success': False,
                'completed': completed,
                'total': total,
                'message': receive_status['message']  # Keep the error message
            }
            receive_status['running'] = False
        elif receive_process.returncode == 0:
            skipped = count_skipped_file()
            receive_status['skipped'] = skipped
            receive_status['current'] = total
            suffix = f' ({skipped} skipped)' if skipped else ''
            receive_status['message'] = f'Completed all {total} items{suffix}'
            receive_status['result'] = {
                'success': True,
                'completed': total,
                'total': total,
                'skipped': skipped,
                'message': f'Completed all {total} receive operations{suffix}'
            }
            receive_status['running'] = False  # Set last to avoid race condition
        elif stderr_text:
            # Real script error with diagnostic output
            receive_status['running'] = False
            receive_status['message'] = f'Script error: {stderr_text}'
            receive_status['result'] = {
                'success': False,
                'completed': receive_status['current'],
                'total': total,
                'message': f'Script error: {stderr_text}'
            }
        else:
            # Non-zero exit with no stderr: process was terminated (user stop
            # or Windows terminate()=1). Clear status silently.
            receive_status['message'] = ''
            receive_status['result'] = {
                'success': False,
                'completed': receive_status['current'],
                'total': total,
                'message': ''
            }
            receive_status['running'] = False
            
    except Exception as e:
        receive_status['running'] = False
        receive_status['message'] = f'Error: {str(e)}'
        receive_status['result'] = {
            'success': False,
            'completed': 0,
            'total': total,
            'message': f'Error: {str(e)}'
        }
    finally:
        # Stop error detector
        if error_detector:
            error_detector.stop()
        receive_process = None
        clear_stop_signal('receive')
        release_device('receive')


@app.route('/execute-receive-batch', methods=['POST'])
def execute_receive_batch():
    """Start batch execution of Receive operations"""
    global receive_status, pending_receive_items
    
    if receive_status['running']:
        return jsonify({'success': False, 'message': 'Another execution is in progress'})

    owner = try_acquire_device('receive')
    if owner:
        return device_busy_response(owner)

    # Use items from request or pending items
    data = request.json or {}
    items = data.get('items', pending_receive_items)
    sublocation = data.get('sublocation', '').strip()

    offset = 0
    total = len(items)
    resumed = apply_resume('receive', receive_status) if data.get('resume') else None
    if resumed:
        offset, total = resumed
    else:
        if not items:
            release_device('receive')
            return jsonify({'success': False, 'message': 'No items to execute'})
        if not sublocation:
            release_device('receive')
            return jsonify({'success': False, 'message': 'Sublocation is required'})

    # Write sublocation to file for receive_typing.py to read. On resume with
    # an empty field (e.g. after a page refresh) the original run's file is kept.
    if sublocation:
        sublocation_file = get_data_file_path('receive_sublocation.txt')
        with open(sublocation_file, 'w') as f:
            f.write(sublocation)

    print(f"Receive batch: sublocation={sublocation or '(unchanged)'}, {total} items"
          f"{f' (resumed, {offset} done)' if resumed else ''}")

    # Reset status
    receive_status = {
        'running': True,
        'current': 0,
        'total': total,
        'offset': offset,
        'current_item': None,
        'message': f'Resuming: {offset} already done...' if resumed else 'Starting batch execution...',
        'result': None,
        'error_detected': False,
        'skipped': count_skipped_file() if resumed else 0,
        'ocr_available': OCR_AVAILABLE
    }
    if not OCR_AVAILABLE:
        print("WARNING: OCR unavailable (Tesseract missing) - duplicate barcodes "
              "will STOP the run instead of being skipped.")

    # Start background thread to run receive_typing.py. A resumed run keeps the
    # skipped-duplicates log accumulating across segments.
    thread = threading.Thread(target=execute_receive_batch_worker,
                              args=(items, total, not resumed))
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'message': f'Started batch: {len(items)} items to {sublocation}'
    })


@app.route('/receive-status-stream')
def receive_status_stream():
    """Server-Sent Events stream for real-time Receive status updates"""
    def generate():
        last_status = None
        while True:
            current = json.dumps(receive_status)
            if current != last_status:
                yield f"data: {current}\n\n"
                last_status = current
            time.sleep(0.3)  # Poll every 300ms
            
            # Stop streaming if not running and we've sent the final status
            if not receive_status['running'] and last_status == current:
                yield f"data: {current}\n\n"
                break
    
    return Response(generate(), mimetype='text/event-stream')


@app.route('/receive-stop', methods=['POST'])
def stop_receive():
    """Stop the current Receive batch execution"""
    global receive_status, receive_process
    
    if not receive_status['running']:
        return jsonify({'success': False, 'message': 'No execution in progress'})
    
    # Terminate the subprocess if running
    if receive_process is not None:
        try:
            receive_process.terminate()
            receive_status['message'] = 'Stopping...'
        except:
            pass
    
    return jsonify({'success': True, 'message': 'Stop requested'})


@app.route('/receive-skipped-count')
def receive_skipped_count():
    """Number of duplicate IMEIs skipped so far (survives page reloads)."""
    return jsonify({'count': count_skipped_file()})


@app.route('/receive-skipped-clear', methods=['POST'])
def clear_receive_skipped():
    """Manually clear the skipped-duplicates log.

    Blocked while a batch runs: the live run counts this file, and clearing
    mid-run would also lose records a Resume is supposed to keep.
    """
    if receive_status['running']:
        return jsonify({'success': False, 'message': 'Cannot clear while a batch is running'})
    path = get_data_file_path('receive_skipped.txt')
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError as e:
        return jsonify({'success': False, 'message': f'Could not clear: {e}'})
    return jsonify({'success': True, 'message': 'Skipped list cleared'})


@app.route('/download-receive-skipped')
def download_receive_skipped():
    """Download the skipped-duplicates log as an Excel file."""
    rows = []
    path = get_data_file_path('receive_skipped.txt')
    if os.path.exists(path):
        with open(path, 'r') as f:
            for line in f:
                if not line.strip():
                    continue
                # imei<TAB>product<TAB>time; old logs have only the IMEI column
                parts = line.rstrip('\n').split('\t')
                rows.append((
                    parts[0].strip(),
                    parts[1].strip() if len(parts) > 1 else '',
                    parts[2].strip() if len(parts) > 2 else ''
                ))

    if not rows:
        return jsonify({'success': False, 'message': 'No skipped IMEIs recorded yet'}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = 'Skipped IMEIs'
    ws.append(['IMEI', 'Product', 'Time skipped'])
    for row in rows:
        ws.append(row)  # values are strings, so IMEIs keep their digits
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"skipped_imeis_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============================================================
# PICK ROUTES
# ============================================================

@app.route('/pick')
def pick_page():
    """Serve the Pick Item interface"""
    return render_template('bulkPick.html')


@app.route('/upload-pick', methods=['POST'])
def upload_pick_excel():
    """
    Handle Excel file upload for Pick.

    Excel format:
    - Column A: IMEIs starting from row 2 (A2)
    - Empty cells trigger ENTER press only
    - Reads down to last non-empty cell in column A
    """
    global pending_pick_items

    # Clear error state and stop signal from any previous run
    pick_status['error_detected'] = False
    clear_stop_signal('pick')

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'File must be .xlsx or .xls'})

    try:
        # Read Excel file
        wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active

        # Pre-load all rows into a list for fast random access (read_only mode)
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        # Find last row with data in column A (1-based indexing)
        last_row = 1
        for idx in range(1, len(all_rows)):  # Start from index 1 = row 2
            row_data = all_rows[idx]
            cell = row_data[0] if row_data and len(row_data) > 0 else None
            if cell is not None and str(cell).strip():
                last_row = idx + 1  # Convert back to 1-based

        if last_row < 2:
            return jsonify({'success': False, 'message': 'No data found in column A starting from row 2'})

        # Read column A from row 2 to last_row (include empty cells)
        items = []
        for idx in range(1, last_row):  # index 1 = row 2, up to last_row
            row_data = all_rows[idx]
            cell = row_data[0] if row_data and len(row_data) > 0 else None
            if cell is not None:
                # Handle numeric IMEIs (Excel may read as float)
                if isinstance(cell, float):
                    cell = int(cell)
                items.append(str(cell).strip())
            else:
                items.append('')  # Empty cell = empty line

        if not items:
            return jsonify({'success': False, 'message': 'No items found in Excel file'})

        # Store items for later execution
        pending_pick_items = items

        # Count non-empty items for display
        non_empty_count = sum(1 for item in items if item.strip())
        empty_count = len(items) - non_empty_count

        return jsonify({
            'success': True,
            'items': items,
            'message': f'Loaded {len(items)} rows ({non_empty_count} IMEIs, {empty_count} empty)'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error reading Excel: {str(e)}'})


def execute_pick_batch_worker(total=None):
    """Background worker for batch Pick execution - runs pick_typing.py"""
    global pick_status, pick_process

    total = total if total is not None else len(pending_pick_items)
    script_path = str(_project_root / 'pick_typing.py')
    progress_file = 'pick_progress.txt'
    error_detector = None
    error_at_item = [None]  # Use list to allow modification in nested function

    def on_error_detected():
        """Callback when error detector finds red screen"""
        # Capture the current progress at the moment of error
        error_at_item[0] = pick_status['current']
        pick_status['error_detected'] = True
        pick_status['message'] = ('Error detected: Red screen - stopping. '
             'Dismiss the error on the device, then press Resume.')
        # Create stop signal FIRST so script sees it before terminate
        create_stop_signal('pick')
        if pick_process:
            try:
                pick_process.terminate()
            except:
                pass

    def on_crash_detected():
        """Callback when app crash/ANR is detected"""
        error_at_item[0] = pick_status['current']
        pick_status['error_detected'] = True
        pick_status['message'] = (
            'App crashed: Finale app stopped responding. '
            'Dialog dismissed and app reopened. Return to the correct page and press Resume.'
        )
        create_stop_signal('pick')
        if pick_process:
            try:
                pick_process.terminate()
            except:
                pass

    try:
        # Clear any existing stop signal before starting
        clear_stop_signal('pick')
        pick_status['current'] = 0
        pick_status['total'] = total
        pick_status['message'] = 'Starting pick_typing.py...'

        # Start error detector
        error_detector = ErrorDetector(
            callback=on_error_detected,
            crash_callback=on_crash_detected
        )
        error_detector.start()

        # Start the process (non-blocking)
        pick_process = subprocess.Popen(
            [sys.executable, script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            cwd=str(_project_root)
        )

        # Poll for progress while process is running
        while pick_process.poll() is None:
            # Check if error was detected - stop updating progress
            if pick_status['error_detected']:
                pick_process.terminate()
                break

            progress = read_progress_file(progress_file)
            if progress:
                current, _ = progress
                # Only update progress if no error detected
                if not pick_status['error_detected']:
                    pick_status['current'] = current
                    pick_status['message'] = f'Processing {current}/{total}...'
            time.sleep(0.3)

        # Process finished
        stdout, stderr = pick_process.communicate()
        stderr_text = stderr.decode().strip() if stderr else ''

        if pick_status.get('error_detected', False):
            # Red screen or crash was detected by ErrorDetector
            completed = error_at_item[0] if error_at_item[0] is not None else pick_status['current']
            pick_status['current'] = completed
            pick_status['result'] = {
                'success': False,
                'completed': completed,
                'total': total,
                'message': pick_status['message']
            }
        elif pick_process.returncode == 0:
            pick_status['current'] = total
            pick_status['message'] = f'Completed {total} items'
            pick_status['result'] = {
                'success': True,
                'completed': total,
                'total': total,
                'message': f'Completed {total} items'
            }
        elif stderr_text:
            # Real script error with diagnostic output
            pick_status['message'] = f'Script error: {stderr_text}'
            pick_status['result'] = {
                'success': False,
                'completed': pick_status['current'],
                'total': total,
                'message': f'Script error: {stderr_text}'
            }
        else:
            # Non-zero exit with no stderr: process was terminated (user stop
            # or Windows terminate()=1). Clear status silently.
            pick_status['message'] = ''
            pick_status['result'] = {
                'success': False,
                'completed': pick_status['current'],
                'total': total,
                'message': ''
            }

    except Exception as e:
        pick_status['message'] = f'Error: {str(e)}'
        pick_status['result'] = {
            'success': False,
            'message': str(e)
        }
    finally:
        pick_status['running'] = False
        if error_detector:
            error_detector.stop()
        pick_process = None
        clear_stop_signal('pick')
        release_device('pick')


@app.route('/execute-pick-batch', methods=['POST'])
def execute_pick_batch():
    """Start batch Pick execution"""
    global pick_status, pending_pick_items

    if pick_status['running']:
        return jsonify({'success': False, 'message': 'Pick batch already in progress'})

    owner = try_acquire_device('pick')
    if owner:
        return device_busy_response(owner)

    data = request.json or {}
    offset = 0
    total = len(pending_pick_items)
    resumed = apply_resume('pick', pick_status) if data.get('resume') else None
    if resumed:
        # Data file already trimmed to the remaining items - don't rewrite it
        offset, total = resumed
    elif not pending_pick_items:
        release_device('pick')
        return jsonify({'success': False, 'message': 'No items loaded. Upload Excel file first.'})
    else:
        # Write to pick_data.txt (preserve empty lines for ENTER-only actions)
        pick_data_file = get_data_file_path('pick_data.txt')
        with open(pick_data_file, 'w') as f:
            for item in pending_pick_items:
                f.write(item + '\n')

    # Reset status
    pick_status['running'] = True
    pick_status['current'] = 0
    pick_status['total'] = total
    pick_status['offset'] = offset
    pick_status['message'] = f'Resuming: {offset} already done...' if resumed else 'Starting...'
    pick_status['error_detected'] = False
    pick_status['result'] = None

    # Start background worker
    thread = threading.Thread(target=execute_pick_batch_worker, args=(total,))
    thread.daemon = True
    thread.start()

    return jsonify({
        'success': True,
        'message': f'Started batch: {len(pending_pick_items)} items'
    })


@app.route('/pick-status-stream')
def pick_status_stream():
    """Server-Sent Events stream for real-time Pick status updates"""
    def generate():
        last_status = None
        while True:
            current = json.dumps(pick_status)
            if current != last_status:
                yield f"data: {current}\n\n"
                last_status = current
            time.sleep(0.3)  # Poll every 300ms

            # Stop streaming if not running and we've sent the final status
            if not pick_status['running'] and last_status == current:
                yield f"data: {current}\n\n"
                break

    return Response(generate(), mimetype='text/event-stream')


@app.route('/pick-stop', methods=['POST'])
def stop_pick():
    """Stop the current Pick batch execution"""
    global pick_status, pick_process

    if not pick_status['running']:
        return jsonify({'success': False, 'message': 'No execution in progress'})

    # Terminate the subprocess if running
    if pick_process is not None:
        try:
            pick_process.terminate()
            pick_status['message'] = 'Stopping...'
        except:
            pass

    return jsonify({'success': True, 'message': 'Stop requested'})


# ============================================================
# RESET ROUTE
# ============================================================

@app.route('/reset', methods=['POST'])
def reset_device():
    """
    Reset all data on the Android device.
    If any process is running, pause and stop it first.
    """
    global execution_status, transfer_status, change_state_status, receive_status, pick_status
    global transfer_process, change_state_process, receive_process, pick_process
    
    stopped_process = None
    
    try:
        # Check and stop any running processes
        
        # Check transfer process
        if transfer_status.get('running'):
            stopped_process = 'Transfer'
            if transfer_process is not None:
                try:
                    transfer_process.terminate()
                except:
                    pass
            transfer_status['running'] = False
            transfer_status['message'] = 'Stopped for reset'
            transfer_status['error_detected'] = False

        # Check change state process
        if change_state_status.get('running'):
            stopped_process = 'Change Item State'
            if change_state_process is not None:
                try:
                    change_state_process.terminate()
                except:
                    pass
            change_state_status['running'] = False
            change_state_status['message'] = 'Stopped for reset'
            change_state_status['error_detected'] = False

        # Check receive process
        if receive_status.get('running'):
            stopped_process = 'Receive'
            if receive_process is not None:
                try:
                    receive_process.terminate()
                except:
                    pass
            receive_status['running'] = False
            receive_status['message'] = 'Stopped for reset'
            receive_status['error_detected'] = False

        # Check pick process
        if pick_status.get('running'):
            stopped_process = 'Pick'
            if pick_process is not None:
                try:
                    pick_process.terminate()
                except:
                    pass
            pick_status['running'] = False
            pick_status['message'] = 'Stopped for reset'
            pick_status['error_detected'] = False

        # Check automator-based execution
        if execution_status.get('running'):
            stopped_process = 'Batch execution'
            auto = get_automator()
            auto.request_stop()
            execution_status['running'] = False
            execution_status['message'] = 'Stopped for reset'
            execution_status['error_detected'] = False
        
        # Small delay to let processes stop
        if stopped_process:
            time.sleep(0.5)

        # Everything is stopped; clear any device claim so flows can start fresh
        force_release_device()

        # Perform the reset
        auto = get_automator()
        result = auto.reset_all_data()
        
        if stopped_process:
            result['message'] = f"Stopped {stopped_process}. {result['message']}"
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': ''
        })


if __name__ == '__main__':
    print("=" * 50)
    print("TRANSFERER SERVER")
    print("Open http://localhost:5000 in your browser")
    print("=" * 50)
    app.run(debug=False, host='127.0.0.1', port=5000, threaded=True)
