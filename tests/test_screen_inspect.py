#!/usr/bin/env python3
"""
Tests for the Receive duplicate-skip pipeline. Run directly (no pytest needed):

    python tests/test_screen_inspect.py

Covers:
  - red-screen detection + OCR classification on synthetic duplicate screens
    (real Tesseract run, both landscape and portrait captures)
  - OCR location of the on-screen Back button
  - product/IMEI line classification (marked + legacy files)
  - skipped-log format and the Excel download / count endpoints
"""
import io
import os
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / 'src'))

from PIL import Image, ImageDraw, ImageFont

import screen_inspect
from adb_utils import get_data_file_path, PRODUCT_MARKER

PASS = 0


def ok(name, condition, detail=""):
    global PASS
    assert condition, f"FAIL: {name} {detail}"
    PASS += 1
    print(f"  ok - {name}")


def _font(size):
    try:
        return ImageFont.truetype(r"C:\Windows\Fonts\arial.ttf", size)
    except Exception:
        return ImageFont.load_default(size)


def make_screen(text, back_label=True, w=2400, h=1080, bg=(200, 30, 30)):
    """Render a synthetic Finale-style error screen (landscape)."""
    img = Image.new("RGB", (w, h), bg)
    draw = ImageDraw.Draw(img)
    draw.text((w // 4, h // 3), text, fill=(255, 255, 255), font=_font(80))
    if back_label:
        draw.text((340, 910), "Back", fill=(255, 255, 255), font=_font(60))
    return img


def test_red_detection():
    print("Red detection:")
    dup = make_screen("This barcode already exists!")
    ok("red screen detected", screen_inspect.is_screen_red(dup))
    clean = Image.new("RGB", (2400, 1080), (245, 245, 245))
    ok("clean screen not red", not screen_inspect.is_screen_red(clean))


def test_classification():
    print("OCR classification (real Tesseract):")
    assert screen_inspect.OCR_AVAILABLE, (
        "OCR_AVAILABLE is False - Tesseract or pytesseract missing; "
        "run start.bat / tools/bootstrap.ps1 first."
    )

    dup = make_screen("This barcode already exists!")
    result = screen_inspect.classify_frame(dup)
    ok("duplicate screen classified", result["state"] == "duplicate", f"got {result}")
    ok("Back button located", result["back_xy"] is not None)
    xv, yv = result["back_xy"]
    # Landscape buffer = 'none' rotation: tap coords == view coords (~340, ~910)
    ok("Back tap coords plausible", 300 <= xv <= 520 and 880 <= yv <= 1010,
       f"got ({xv}, {yv})")

    # Portrait capture, clockwise case: rotating the buffer CW recovers the
    # view, so tap = (yv, bufH-1-xv). This is the live-verified device mapping.
    portrait_cw = make_screen("This barcode already exists!").rotate(90, expand=True)
    r_cw = screen_inspect.classify_frame(portrait_cw)
    ok("portrait(cw) classified", r_cw["state"] == "duplicate", f"got {r_cw}")
    ex, ey = yv, 2400 - 1 - xv
    ok("portrait(cw) tap mapped to buffer coords",
       abs(r_cw["back_xy"][0] - ex) <= 3 and abs(r_cw["back_xy"][1] - ey) <= 3,
       f"expected ~({ex},{ey}) got {r_cw['back_xy']}")

    # Portrait capture, counter-clockwise case: tap = (bufW-1-yv, xv)
    portrait_ccw = make_screen("This barcode already exists!").rotate(-90, expand=True)
    r_ccw = screen_inspect.classify_frame(portrait_ccw)
    ok("portrait(ccw) classified", r_ccw["state"] == "duplicate", f"got {r_ccw}")
    ex2, ey2 = 1080 - 1 - yv, xv
    ok("portrait(ccw) tap mapped to buffer coords",
       abs(r_ccw["back_xy"][0] - ex2) <= 3 and abs(r_ccw["back_xy"][1] - ey2) <= 3,
       f"expected ~({ex2},{ey2}) got {r_ccw['back_xy']}")

    # Duplicate confirmed but no 'Back' word found -> live-verified fallback
    dup_no_back = make_screen("This barcode already exists!", back_label=False)
    r_fb = screen_inspect.classify_frame(dup_no_back)
    ok("fallback to verified Back position",
       r_fb["state"] == "duplicate" and r_fb["back_xy"] == screen_inspect.BACK_TAP_FALLBACK,
       f"got {r_fb}")

    other = make_screen("Network connection failure", back_label=False)
    result_o = screen_inspect.classify_frame(other)
    ok("non-duplicate red is other_red", result_o["state"] == "other_red", f"got {result_o}")


def test_parse_receive_items():
    print("parse_receive_items:")
    from receive_typing import parse_receive_items

    marked = [
        f"{PRODUCT_MARKER}PIXEL 8 128GB BLACK",
        "356789012345678",
        "356789012345679",
        f"{PRODUCT_MARKER}IPAD MINI 4 WI-FI 128GB",
        "FAKE0SERIAL1",
    ]
    items = parse_receive_items(marked)
    ok("marked: kinds", [k for k, _ in items] == ["product", "imei", "imei", "product", "imei"])
    ok("marked: marker stripped", items[0][1] == "PIXEL 8 128GB BLACK")
    ok("marked: non-Apple product recognized", items[0][0] == "product")

    legacy = [
        "IPAD MINI 4 WI-FI 128GB GOLD A1538",
        "FAKE0SERIAL1",
        "356789012345678",
    ]
    items_l = parse_receive_items(legacy)
    ok("legacy: prefix fallback", [k for k, _ in items_l] == ["product", "imei", "imei"])


def test_skipped_log_and_excel():
    print("Skipped log + Excel download:")
    from openpyxl import load_workbook
    import receive_typing

    skipped_path = get_data_file_path("receive_skipped.txt")
    backup = None
    if os.path.exists(skipped_path):
        with open(skipped_path, "r") as f:
            backup = f.read()
    try:
        if os.path.exists(skipped_path):
            os.remove(skipped_path)

        receive_typing.append_skipped("356789012345678", "PIXEL 8 128GB BLACK")
        receive_typing.append_skipped("356789012345679")  # product unknown

        from transferer_server import app
        client = app.test_client()

        count = client.get("/receive-skipped-count").get_json()
        ok("count endpoint", count["count"] == 2, f"got {count}")

        resp = client.get("/download-receive-skipped")
        ok("download status 200", resp.status_code == 200, f"got {resp.status_code}")
        ok("download is xlsx", "spreadsheetml" in resp.headers.get("Content-Type", ""))

        wb = load_workbook(io.BytesIO(resp.data))
        rows = list(wb.active.iter_rows(values_only=True))
        ok("xlsx header", rows[0] == ("IMEI", "Product", "Time skipped"))
        ok("xlsx row 1", rows[1][0] == "356789012345678" and rows[1][1] == "PIXEL 8 128GB BLACK")
        ok("xlsx row 2", rows[2][0] == "356789012345679" and rows[2][1] in ("", None))
        ok("IMEI stays text", isinstance(rows[1][0], str))
        ok("timestamp present", rows[1][2] and rows[1][2][:2] == "20")

        # Empty log -> 404, count 0
        os.remove(skipped_path)
        ok("empty count", client.get("/receive-skipped-count").get_json()["count"] == 0)
        ok("empty download 404", client.get("/download-receive-skipped").status_code == 404)
    finally:
        if backup is not None:
            with open(skipped_path, "w") as f:
                f.write(backup)
        elif os.path.exists(skipped_path):
            os.remove(skipped_path)


def test_read_screen_state_unknown(monkey_adb="definitely-not-adb.exe"):
    print("read_screen_state fail-closed:")
    # A bogus adb path means every capture fails -> must report 'unknown',
    # never 'clear' (which would let the script type blindly).
    state = screen_inspect.read_screen_state(monkey_adb, attempts=2, interval=0.01)
    ok("all captures failed -> unknown", state["state"] == "unknown", f"got {state}")


if __name__ == "__main__":
    test_red_detection()
    test_classification()
    test_parse_receive_items()
    test_skipped_log_and_excel()
    test_read_screen_state_unknown()
    print(f"\nALL {PASS} CHECKS PASSED")
